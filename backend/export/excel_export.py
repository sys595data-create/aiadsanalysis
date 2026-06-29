import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")


def _write_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def build_excel(
    raw_items: list[dict],
    clusters: dict,
    triaged: list[dict],
    recipes: list[dict],
    stats: dict,
    voc_analytics: dict,
    ad_analytics: dict,
    run_id: str,
) -> bytes:
    wb = Workbook()

    # Sheet 1 — README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "aiadsanalysis — Market Intelligence Run"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Run ID: {run_id}"
    ws["A3"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A5"] = f"Total items: {len(raw_items)}"
    ws["A6"] = f"Content clusters: {len(clusters.get('content_clusters', []))}"
    ws["A7"] = f"VOC clusters: {len(clusters.get('voc_clusters', []))}"
    ws["A8"] = f"Ad clusters: {len(clusters.get('ad_clusters', []))}"
    ws["A9"] = f"Triaged for video processing: {len(triaged)}"
    ws["A11"] = "Sheets: Raw Content | Clusters | VOC | Ads | Recipes | Statistics | VOC Analytics | Ad Analytics | Home Demo"

    # Sheet 2 — Raw Content
    ws2 = wb.create_sheet("Raw Content")
    headers2 = ["Source", "Market", "Layer", "Category", "Type", "Brand", "Product", "Text (truncated)", "Views", "Likes", "Engagement", "URL", "Date"]
    _write_header(ws2, headers2)
    for i, item in enumerate(raw_items, 2):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [
            item.get("source"), item.get("market"), item.get("layer"),
            item.get("data_category"), item.get("content_type"),
            item.get("brand"), item.get("product"),
            (item.get("text") or "")[:300],
            item.get("view_count"), item.get("like_count"),
            round(item.get("engagement_score", 0), 4),
            item.get("url"), item.get("date_str"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            if fill:
                cell.fill = fill
    _autowidth(ws2)

    # Sheet 3 — Content Clusters
    ws3 = wb.create_sheet("Clusters")
    headers3 = ["ID", "Type", "Keywords", "Size", "Sentiment", "Layer Mix", "Sample 1", "Sample 2"]
    _write_header(ws3, headers3)
    all_clusters = (
        clusters.get("content_clusters", []) +
        clusters.get("voc_clusters", []) +
        clusters.get("ad_clusters", [])
    )
    for i, c in enumerate(all_clusters, 2):
        ws3.append([
            c.get("id"), c.get("type"),
            ", ".join(c.get("keywords", [])[:8]),
            c.get("size"), round(c.get("sentiment", 0), 3),
            str(c.get("layer_mix", {})),
            (c.get("sample_texts") or [""])[0][:200],
            (c.get("sample_texts") or ["", ""])[1][:200] if len(c.get("sample_texts", [])) > 1 else "",
        ])
    _autowidth(ws3)

    # Sheet 4 — VOC
    ws4 = wb.create_sheet("VOC")
    headers4 = ["Pain", "Intensity", "Emotional State", "Failed Solution", "Opportunity", "Hook"]
    _write_header(ws4, headers4)
    pains = voc_analytics.get("primary_pains", [])
    failed = voc_analytics.get("failed_solutions", [])
    opps = voc_analytics.get("opportunities", [])
    emotions = voc_analytics.get("emotional_states", [])
    for i, pain in enumerate(pains, 2):
        ws4.append([
            pain.get("pain"), pain.get("intensity"),
            emotions[i - 2] if i - 2 < len(emotions) else "",
            failed[i - 2] if i - 2 < len(failed) else "",
            opps[i - 2].get("gap", "") if i - 2 < len(opps) else "",
            opps[i - 2].get("hook", "") if i - 2 < len(opps) else "",
        ])
    _autowidth(ws4)

    # Sheet 5 — Ad Analytics
    ws5 = wb.create_sheet("Ad Analytics")
    headers5 = ["Hook", "Why It Works", "Audience Segment", "USP", "Missing Angle"]
    _write_header(ws5, headers5)
    hooks = ad_analytics.get("top_hooks", [])
    segments = ad_analytics.get("audience_segments", [])
    usps = ad_analytics.get("usp_positions", [])
    gaps = ad_analytics.get("missing_angles", [])
    max_rows = max(len(hooks), len(segments), len(usps), len(gaps), 1)
    for i in range(max_rows):
        ws5.append([
            hooks[i].get("hook", "") if i < len(hooks) else "",
            hooks[i].get("why_it_works", "") if i < len(hooks) else "",
            segments[i].get("segment", "") if i < len(segments) else "",
            usps[i].get("usp", "") if i < len(usps) else "",
            gaps[i] if i < len(gaps) else "",
        ])
    _autowidth(ws5)

    # Sheet 6 — Recipes
    ws6 = wb.create_sheet("Recipes")
    headers6 = ["Cluster ID", "Market", "Hook Type", "Opening Line", "Setting", "Engagement Prediction", "Competitor Gaps", "Script Outline"]
    _write_header(ws6, headers6)
    for recipe in recipes:
        hook = recipe.get("hook_architecture", {})
        ws6.append([
            recipe.get("cluster_id"),
            recipe.get("market", "us"),
            hook.get("hook_type", ""),
            hook.get("opening_line", ""),
            recipe.get("setting", {}).get("type", ""),
            recipe.get("engagement_prediction", ""),
            "; ".join(recipe.get("competitor_gaps", [])),
            str(recipe.get("script_outline", [])),
        ])
    _autowidth(ws6)

    # Sheet 7 — Statistics
    ws7 = wb.create_sheet("Statistics")
    ws7["A1"] = "Statistics"
    ws7["A1"].font = Font(bold=True)
    row = 2
    for k, v in stats.items():
        ws7.cell(row=row, column=1, value=str(k))
        ws7.cell(row=row, column=2, value=str(v))
        row += 1
    _autowidth(ws7)

    # Sheet 8 — Triaged Items
    ws8 = wb.create_sheet("Triaged Videos")
    headers8 = ["Source", "Market", "Layer", "Engagement", "Transcript (truncated)", "Visual Concepts", "URL"]
    _write_header(ws8, headers8)
    for item in triaged:
        ws8.append([
            item.get("source"), item.get("market"), item.get("layer"),
            round(item.get("engagement_score", 0), 4),
            (item.get("transcript") or "")[:400],
            ", ".join(item.get("visual_concepts", [])),
            item.get("url"),
        ])
    _autowidth(ws8)

    # Sheet 9 — Word Cloud Terms
    ws9 = wb.create_sheet("VOC Word Cloud")
    headers9 = ["Term", "Frequency"]
    _write_header(ws9, headers9)
    wc = voc_analytics.get("word_cloud_terms", {})
    for term, freq in sorted(wc.items(), key=lambda x: -x[1])[:100]:
        ws9.append([term, freq])
    _autowidth(ws9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
